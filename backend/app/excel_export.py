from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import tempfile
import os
from datetime import datetime

def create_excel_report(result: dict, site_id: int) -> str:
    """
    Создание Excel отчета с результатами проверки фида
    
    Returns:
        str: Путь к созданному файлу
    """
    wb = Workbook()
    
    # Удаляем дефолтный лист
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 1. Общая статистика
    ws_summary = wb.create_sheet("Общая статистика")
    _create_summary_sheet(ws_summary, result, site_id)
    
    # 2. Проблемные товары
    ws_problems = wb.create_sheet("Проблемные товары")
    _create_problems_sheet(ws_problems, result)
    
    # 3. Проблемные категории
    ws_categories = wb.create_sheet("Проблемные категории")
    _create_categories_sheet(ws_categories, result)
    
    # 4. Дубликаты ID
    if result.get("mandatory", {}).get("duplicate_ids"):
        ws_duplicates = wb.create_sheet("Дубликаты ID")
        _create_duplicates_sheet(ws_duplicates, result)
    
    # Сохранение файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"feed_check_{site_id}_{timestamp}.xlsx"
    temp_dir = tempfile.gettempdir()
    filepath = os.path.join(temp_dir, filename)
    
    wb.save(filepath)
    return filepath

def _create_summary_sheet(ws, result: dict, site_id: int):
    """Создание листа с общей статистикой"""
    # Заголовок
    ws.append(["FeedChecker - Отчет по проверке фида"])
    ws.append([])
    ws.append([f"Site ID: {site_id}"])
    ws.append([f"Дата проверки: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"])
    ws.append([])
    
    # Стилизация заголовка
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')
    
    mandatory = result.get("mandatory", {})
    
    # Общая информация
    ws.append(["ОБЩАЯ ИНФОРМАЦИЯ", ""])
    ws['A6'].font = Font(size=14, bold=True)
    ws['A6'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A6'].font = Font(color="FFFFFF", bold=True)
    
    ws.append(["Всего товаров:", mandatory.get("total_offers", 0)])
    ws.append(["Доступных товаров:", mandatory.get("available_offers", 0)])
    ws.append(["Недоступных товаров:", mandatory.get("unavailable_offers", 0)])
    ws.append(["Всего категорий:", mandatory.get("total_categories", 0)])
    ws.append(["Глубина дерева категорий:", mandatory.get("category_tree_depth", 0)])
    ws.append(["Количество брендов:", mandatory.get("brands_count", 0)])
    ws.append([])
    
    # Проблемы
    ws.append(["ОБНАРУЖЕННЫЕ ПРОБЛЕМЫ", ""])
    ws[f'A{ws.max_row}'].font = Font(size=14, bold=True)
    ws[f'A{ws.max_row}'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    ws[f'A{ws.max_row}'].font = Font(color="FFFFFF", bold=True)
    
    problems = mandatory.get("problems", {})
    problem_labels = {
        "missing_id": "Без ID",
        "missing_availability": "Без информации о доступности",
        "missing_name": "Без названия",
        "missing_link": "Без ссылки",
        "price_issues": "Проблемы с ценой",
        "missing_category": "Без категории",
        "invalid_category": "Недействительная категория",
        "multiple_categories": "Несколько категорий без тега",
        "vendor_issues": "Проблемы с брендом",
        "missing_image": "Без изображения",
    }
    
    for key, count in problems.items():
        label = problem_labels.get(key, key)
        ws.append([label, count])
        if count > 0:
            ws[f'B{ws.max_row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Автоширина колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15

def _create_problems_sheet(ws, result: dict):
    """Создание листа с проблемными товарами с детальной информацией"""
    # Заголовки
    ws.append(["Тип проблемы", "ID товара", "Название", "Бренд", "Категории", "Цена", "URL"])

    # Стилизация заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    problem_labels = {
        "missing_id": "Без ID",
        "missing_availability": "Без информации о доступности",
        "missing_name": "Без названия",
        "missing_link": "Без ссылки",
        "price_issues": "Проблемы с ценой",
        "missing_category": "Без категории",
        "invalid_category": "Недействительная категория",
        "multiple_categories": "Несколько категорий без тега",
        "vendor_issues": "Проблемы с брендом",
        "missing_image": "Без изображения",
    }

    # Получаем детальную информацию о проблемных товарах
    problematic_offers = result.get("problematic_offers", {})

    # Для каждого типа проблемы выводим все товары
    for key, label in problem_labels.items():
        offers = problematic_offers.get(key, [])

        if offers:
            # Добавляем товары с этой проблемой
            for offer in offers:
                ws.append([
                    label,
                    offer.get("id", "N/A"),
                    offer.get("name", "N/A"),
                    offer.get("vendor", "N/A"),
                    offer.get("categories", "N/A"),
                    offer.get("price", "N/A"),
                    offer.get("url", "N/A")
                ])
                # Подсвечиваем строки с проблемами
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    ws[f'{col}{ws.max_row}'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")

    # Автоширина колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 60

def _create_categories_sheet(ws, result: dict):
    """Создание листа с проблемными категориями"""
    categories = result.get("categories", {})
    
    # Заголовки
    ws.append(["ID категории", "Название", "Тип проблемы", "Влияние"])
    
    # Стилизация заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Пустые категории
    for cat_id, cat_name in categories.get("empty_categories", []):
        ws.append([cat_id, cat_name, "Пустая категория", "Не влияет на поиск"])
    
    # Дубликаты
    for cat_id, cat_name in categories.get("duplicated_categories", []):
        ws.append([cat_id, cat_name, "Дубликат", "Влияет на подсказки"])
        ws[f'A{ws.max_row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Сдвоенные
    for cat_id, cat_name in categories.get("dual_categories", []):
        ws.append([cat_id, cat_name, "Сдвоенная категория", "Влияет на отбор кандидатов"])
        ws[f'A{ws.max_row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Автоширина колонок
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30

def _create_duplicates_sheet(ws, result: dict):
    """Создание листа с дубликатами ID"""
    # Заголовки
    ws.append(["ID товара", "Количество повторений"])
    
    # Стилизация заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    duplicate_ids = result.get("mandatory", {}).get("duplicate_ids", [])
    
    for offer_id, count in duplicate_ids:
        ws.append([offer_id, count])
        ws[f'A{ws.max_row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Автоширина колонок
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25

